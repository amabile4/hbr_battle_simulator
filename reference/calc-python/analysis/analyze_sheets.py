import os
import json
import re
from openpyxl import load_workbook

# パス定義
EXCEL_PATH = "HBR計算機\U0001f3adVer.4.31.03.xlsx"
OUTPUT_DIR = "extracted_info"
os.makedirs(OUTPUT_DIR, exist_ok=True)

def analyze_workbook():
    print(f"Loading workbook: {EXCEL_PATH} ...")
    # data_only=False で数式を取得
    wb = load_workbook(EXCEL_PATH, data_only=False, read_only=False)
    
    summary = {
        "sheets": [],
        "defined_names": [],
        "function_usage": {}
    }
    
    # 1. 定義された名前 (Named Ranges) の抽出
    print("Extracting defined names...")
    for name, def_name in wb.defined_names.items():
        # def_name.value または def_name.destinations で参照先を取得
        destinations = []
        try:
            for title, coord in def_name.destinations:
                destinations.append({"sheet": title, "coordinate": coord})
        except Exception:
            destinations = str(def_name.value)
            
        summary["defined_names"].append({
            "name": name,
            "destinations": destinations,
            "local_only": def_name.localSheetId is not None,
            "comment": def_name.comment
        })
        
    # Excel関数を抽出するための正規表現 (例: VLOOKUP, INDEX, IF など)
    # 一般的に英大文字で始まり、直後に括弧が来るパターン
    func_pattern = re.compile(r'\b([A-Z_]+)\s*\(')
    
    formulas_data = {}
    
    # 2. 各シートの解析
    for sheet_name in wb.sheetnames:
        print(f"Analyzing sheet: {sheet_name} ...")
        ws = wb[sheet_name]
        
        # シートの基本情報
        sheet_info = {
            "name": sheet_name,
            "state": ws.sheet_state, # visible, hidden, veryHidden
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "formula_cells_count": 0,
            "total_cells_count": 0
        }
        
        sheet_formulas = []
        
        # 全セルを走査して数式を抽出
        # メモリ削減のため、値があるセルのみ走査
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if cell.value is not None:
                    sheet_info["total_cells_count"] += 1
                    val_str = str(cell.value)
                    if val_str.startswith('='):
                        sheet_info["formula_cells_count"] += 1
                        
                        # 数式情報の保存
                        sheet_formulas.append({
                            "cell": cell.coordinate,
                            "formula": val_str
                        })
                        
                        # 使用されている関数の抽出
                        funcs = func_pattern.findall(val_str)
                        for func in funcs:
                            summary["function_usage"][func] = summary["function_usage"].get(func, 0) + 1
                            
        summary["sheets"].append(sheet_info)
        formulas_data[sheet_name] = sheet_formulas
        
    # 結果の書き出し
    summary_path = os.path.join(OUTPUT_DIR, "workbook_summary.json")
    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
        
    formulas_path = os.path.join(OUTPUT_DIR, "formulas_by_sheet.json")
    with open(formulas_path, "w", encoding="utf-8") as f:
        json.dump(formulas_data, f, ensure_ascii=False, indent=2)
        
    # テキスト形式でのサマリーレポートも作成
    report_path = os.path.join(OUTPUT_DIR, "analysis_report.txt")
    with open(report_path, "w", encoding="utf-8") as f:
        f.write("=== HBR計算機 構造解析レポート ===\n\n")
        f.write(f"解析ファイル: {EXCEL_PATH}\n")
        f.write(f"総シート数: {len(wb.sheetnames)}\n\n")
        
        f.write("--- シート一覧 ---\n")
        for s in summary["sheets"]:
            f.write(f"シート名: {s['name']}\n")
            f.write(f"  状態: {s['state']}\n")
            f.write(f"  サイズ: {s['max_row']}行 x {s['max_column']}列\n")
            f.write(f"  総セル数(非空): {s['total_cells_count']}\n")
            f.write(f"  数式セル数: {s['formula_cells_count']}\n\n")
            
        f.write("--- 使用されているExcel関数 (頻度順) ---\n")
        sorted_funcs = sorted(summary["function_usage"].items(), key=lambda x: x[1], reverse=True)
        for func, count in sorted_funcs:
            f.write(f"  {func}: {count}回\n")
        f.write("\n")
        
        f.write("--- 定義された名前 (Named Ranges) 一覧 ---\n")
        for dn in summary["defined_names"]:
            f.write(f"名前: {dn['name']}\n")
            f.write(f"  参照先: {dn['destinations']}\n")
            f.write(f"  ローカルスコープ: {dn['local_only']}\n\n")

    print("Analysis completed successfully.")
    print(f"Results saved to: {OUTPUT_DIR}/")

if __name__ == "__main__":
    analyze_workbook()
