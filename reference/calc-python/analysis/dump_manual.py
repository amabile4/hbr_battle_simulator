from openpyxl import load_workbook
import os

EXCEL_PATH = "HBR計算機\U0001f3adVer.4.31.03.xlsx"

def dump_manual():
    wb = load_workbook(EXCEL_PATH, data_only=True)
    if "マニュアル" not in wb.sheetnames:
        print("マニュアル sheet not found.")
        return
        
    ws = wb["マニュアル"]
    print("Dumping マニュアル sheet...")
    
    lines = []
    for r in range(1, ws.max_row + 1):
        row_vals = []
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                row_vals.append(f"Col{c}: {str(val).strip()}")
        if row_vals:
            lines.append(f"Row {r}: " + " | ".join(row_vals))
            
    output_path = "extracted_info/manual_dump.txt"
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"Dumped to {output_path}")

if __name__ == "__main__":
    dump_manual()
