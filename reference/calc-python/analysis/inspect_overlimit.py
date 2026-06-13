import openpyxl

EXCEL_PATH = "data/HBR計算機🎭Ver.4.31.03_custom_fixed.xlsx"

def inspect():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
    ws = wb["ダメージ計算機"]
    
    cells = ["W59", "X59", "R59", "P59", "N59", "M59", "U59", "AJ8", "AY6"]
    print("=== Formula Inspection ===")
    for cell_coord in cells:
        cell = ws[cell_coord]
        print(f"Cell {cell_coord}: {cell.value}")

    wb_val = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws_val = wb_val["ダメージ計算機"]
    print("\n=== Current Value Inspection ===")
    for cell_coord in cells:
        cell = ws_val[cell_coord]
        print(f"Cell {cell_coord}: {cell.value}")

if __name__ == "__main__":
    inspect()
