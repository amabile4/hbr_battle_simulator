import openpyxl

EXCEL_PATH = "data/HBR計算機🎭Ver.4.31.03_custom_fixed.xlsx"

def inspect():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
    ws = wb["ダメージ計算機"]
    
    wb_val = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws_val = wb_val["ダメージ計算機"]
    
    row = 59
    print(f"=== Row {row} Inspection ===")
    for col in range(1, 40): # A to AM
        col_letter = openpyxl.utils.get_column_letter(col)
        cell_f = ws.cell(row=row, column=col)
        cell_v = ws_val.cell(row=row, column=col)
        # Find cell label if any (usually row 3 or 4 or 5)
        label1 = ws_val.cell(row=3, column=col).value
        label2 = ws_val.cell(row=4, column=col).value
        label = f"{label1} | {label2}" if label1 or label2 else ""
        print(f"Col {col_letter} ({label}): val={cell_v.value} | formula={cell_f.value}")

if __name__ == "__main__":
    inspect()
