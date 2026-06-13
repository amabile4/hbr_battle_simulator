from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import json

EXCEL_PATH = "HBR計算機\U0001f3adVer.4.31.03.xlsx"
OUTPUT_DIR = "extracted_info"
os.makedirs(OUTPUT_DIR, exist_ok=True)

def analyze_cells():
    wb = load_workbook(EXCEL_PATH, data_only=False)
    ws = wb["ダメージ計算機"]
    
    print(f"Scanning sheet 'ダメージ計算機' ({ws.max_row} rows x {ws.max_column} cols)...")
    
    # 色の出現頻度と代表セルの記録
    color_map = {} # argb -> list of coords
    data_validation_cells = []
    
    # データの入力規則 (Data Validation) のスキャン
    # openpyxl では ws.data_validations.dataValidation で取得可能
    for dv in ws.data_validations.dataValidation:
        # dv.sqref にセルの範囲（例: "A1:B2" や "C5"）が入っている
        sqref_str = str(dv.sqref)
        data_validation_cells.append({
            "formula": dv.formula1,
            "type": dv.type,
            "sqref": sqref_str
        })
        
    cell_details = []
    
    # 行と列を限定してスキャン (データがある範囲のみ)
    # あまりに広いと遅いため、行は1〜150、列は1〜100 (A〜CV程度) を走査
    max_r = min(150, ws.max_row)
    max_c = min(100, ws.max_column)
    
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            cell = ws.cell(row=r, column=c)
            coord = cell.coordinate
            val = cell.value
            
            # 背景色取得
            color_hex = "FFFFFF"
            if cell.fill and isinstance(cell.fill, PatternFill) and cell.fill.start_color:
                # start_color.rgb は "00000000" (ARGB) またはインデックスの可能性あり
                rgb = cell.fill.start_color.rgb
                if rgb and isinstance(rgb, str) and rgb != "00000000":
                    color_hex = rgb
            
            if color_hex not in color_map:
                color_map[color_hex] = []
            color_map[color_hex].append(coord)
            
            # 特徴的なセルの抽出 (値または数式があるもの)
            if val is not None:
                val_str = str(val)
                is_formula = val_str.startswith("=")
                
                # フォント情報
                is_bold = False
                font_color = "000000"
                if cell.font:
                    is_bold = bool(cell.font.bold)
                    if cell.font.color and cell.font.color.rgb:
                        font_color = str(cell.font.color.rgb)
                
                cell_details.append({
                    "cell": coord,
                    "value": val_str,
                    "is_formula": is_formula,
                    "color": color_hex,
                    "bold": is_bold,
                    "font_color": font_color
                })
                
    # 統計出力
    report = {
        "color_summary": {k: {"count": len(v), "sample": v[:5]} for k, v in color_map.items() if k != "FFFFFF"},
        "data_validations": data_validation_cells,
        "sample_cells": cell_details[:200] # 一部サンプルの詳細
    }
    
    with open(os.path.join(OUTPUT_DIR, "damage_calc_cells.json"), "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
        
    print("Analysis finished. Saved to damage_calc_cells.json")

if __name__ == "__main__":
    analyze_cells()
