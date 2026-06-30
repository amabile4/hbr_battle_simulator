import openpyxl

EXCEL_PATH = "data/HBR計算機🎭Ver.4.31.03_custom_fixed.xlsx"

def search():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    keywords = ["超過", "上限", "0.25", "0.0025", "頭打ち", "限界"]
    
    print("Searching Excel workbook sheets...")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Skip very large hidden sheets if we just want a quick search, but let's scan all
        for r in range(1, min(200, ws.max_row) + 1):
            for c in range(1, min(100, ws.max_column) + 1):
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    val_str = str(val)
                    for kw in keywords:
                        if kw in val_str:
                            print(f"[{sheet_name}] Cell {openpyxl.utils.get_column_letter(c)}{r}: found '{kw}' in '{val_str[:50]}'")

if __name__ == "__main__":
    search()
