import openpyxl

EXCEL_PATH = "HBR計算機\U0001f3adVer.4.31.03.xlsx"

def dump_stats():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["能力値"]
    
    print(f"Stats sheet shape: {ws.max_row} rows x {ws.max_column} cols")
    
    # 最初の30行のうち、値があるセルを出力
    for r in range(1, 120):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 10)]
        if any(v is not None for v in row_vals):
            print(f"Row {r:03d}: {row_vals}")

if __name__ == "__main__":
    dump_stats()
