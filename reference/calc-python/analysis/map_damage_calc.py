from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

EXCEL_PATH = "HBR計算機\U0001f3adVer.4.31.03.xlsx"
OUTPUT_PATH = "extracted_info/damage_calc_layout.txt"

def analyze_layout():
    wb = load_workbook(EXCEL_PATH, data_only=True) # 値をロード
    ws = wb["ダメージ計算機"]
    
    wb_formula = load_workbook(EXCEL_PATH, data_only=False) # 数式をロード
    ws_formula = wb_formula["ダメージ計算機"]
    
    # 走査範囲の設定
    # 一般的な入力シートは行100、列100以内にあると想定
    max_r = min(120, ws.max_row)
    max_c = min(80, ws.max_column)
    
    # セル値の分類用キーワード
    keywords = ["ダメージ", "攻撃力", "知性", "精神", "器用さ", "体力", "運", "バフ", "デバフ", "敵", "防御", "耐性", "スキル", "クリティカル"]
    
    print("Mapping spreadsheet UI layout...")
    
    # 行ごとに情報を集約
    layout_grid = []
    
    # 色情報を一時ストック
    colors = {}
    
    for r in range(1, max_r + 1):
        row_cells = []
        for c in range(1, max_c + 1):
            cell = ws.cell(row=r, column=c)
            cell_formula = ws_formula.cell(row=r, column=c)
            coord = cell.coordinate
            val = cell.value
            formula = cell_formula.value
            
            # 背景色
            color = "FFFFFF"
            if cell.fill and isinstance(cell.fill, PatternFill) and cell.fill.start_color:
                rgb = cell.fill.start_color.rgb
                if rgb and isinstance(rgb, str) and rgb != "00000000":
                    color = rgb
            
            if color != "FFFFFF":
                colors[color] = colors.get(color, 0) + 1
                
            # 入力規則 (Data Validation) の判定
            has_dv = False
            for dv in ws.data_validations.dataValidation:
                if coord in dv:
                    has_dv = True
                    break
            
            # セルの分類
            cell_type = "EMPTY"
            if val is not None:
                val_str = str(val).strip()
                is_formula = str(formula).startswith("=")
                
                if is_formula:
                    cell_type = "FORMULA"
                elif has_dv:
                    cell_type = "INPUT_LIST"
                elif val_str.replace(".", "").replace("-", "").isdigit():
                    cell_type = "NUMBER"
                else:
                    cell_type = "LABEL"
                    
            row_cells.append({
                "coord": coord,
                "val": val,
                "formula": formula,
                "type": cell_type,
                "color": color,
                "has_dv": has_dv
            })
        layout_grid.append(row_cells)
        
    # 分析レポートの作成
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write("=== ダメージ計算機 レイアウト分析 ===\n\n")
        
        # 1. 色の統計と仮定される役割
        f.write("--- 背景色と出現数 (上位) ---\n")
        sorted_colors = sorted(colors.items(), key=lambda x: x[1], reverse=True)
        for color, count in sorted_colors:
            # その色を持つ代表的なセルとその値を探す
            samples = []
            for r in range(len(layout_grid)):
                for c in range(len(layout_grid[r])):
                    cell = layout_grid[r][c]
                    if cell["color"] == color and cell["val"] is not None:
                        samples.append(f"{cell['coord']}='{cell['val']}'({cell['type']})")
                        if len(samples) >= 3:
                            break
                if len(samples) >= 3:
                    break
            f.write(f"Color #{color}: {count} cells. Samples: {', '.join(samples)}\n")
        f.write("\n")
        
        # 2. 主要なインプット項目 (入力規則があるセル)
        f.write("--- 主要入力セル (リスト選択/プルダウン) ---\n")
        for r in range(len(layout_grid)):
            for c in range(len(layout_grid[r])):
                cell = layout_grid[r][c]
                if cell["has_dv"]:
                    f.write(f"Cell {cell['coord']}: Value='{cell['val']}', Formula='{cell['formula']}' (Type: {cell['type']})\n")
        f.write("\n")
        
        # 3. ダメージに関連するキーワードを含むラベルとその周辺の数式 (OUTの特定)
        f.write("--- 『ダメージ』『合計』などの重要ラベルと周辺セル ---\n")
        for r in range(len(layout_grid)):
            for c in range(len(layout_grid[r])):
                cell = layout_grid[r][c]
                if cell["type"] == "LABEL" and any(k in str(cell["val"]) for k in ["ダメージ", "合計", "対DP", "対HP", "計算結果"]):
                    f.write(f"Label Cell {cell['coord']}: '{cell['val']}'\n")
                    # 周辺セル (右または下) を出力
                    for offset_r in [-1, 0, 1]:
                        for offset_c in [-1, 0, 1, 2, 3]:
                            nr = r + offset_r
                            nc = c + offset_c
                            if 0 <= nr < len(layout_grid) and 0 <= nc < len(layout_grid[nr]):
                                neighbor = layout_grid[nr][nc]
                                if neighbor["coord"] != cell["coord"] and neighbor["val"] is not None:
                                    f.write(f"  Neighbor {neighbor['coord']} ({offset_r},{offset_c}): Value='{neighbor['val']}', Formula='{neighbor['formula']}'\n")
                    f.write("-" * 40 + "\n")
                    
        # 4. 横方向の簡易レイアウトビジュアルマップ (大雑把な配置の把握)
        f.write("\n--- シート全体の簡易UIレイアウトマップ (A〜BC列) ---\n")
        f.write("凡例: L=LABEL, I=INPUT_LIST, N=NUMBER, F=FORMULA, .=EMPTY\n\n")
        
        # ヘッダー列名
        col_headers = [get_column_letter(c) for c in range(1, max_c + 1)]
        f.write("    " + "".join([ch.ljust(3) for ch in col_headers if len(ch) == 1] + [ch.ljust(3) for ch in col_headers if len(ch) > 1]) + "\n")
        
        for r_idx, row in enumerate(layout_grid):
            row_str = f"{r_idx+1:03d} "
            for cell in row:
                t = cell["type"]
                char = "."
                if t == "LABEL": char = "L"
                elif t == "INPUT_LIST": char = "I"
                elif t == "NUMBER": char = "N"
                elif t == "FORMULA": char = "F"
                
                # 色がついている場合は大文字、白い場合は小文字にするなどの区別
                if cell["color"] != "FFFFFF" and char != ".":
                    char = char.upper()
                elif char != ".":
                    char = char.lower()
                    
                row_str += char.ljust(3)
            f.write(row_str + "\n")

def get_column_letter(col_idx):
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result

if __name__ == "__main__":
    analyze_layout()
