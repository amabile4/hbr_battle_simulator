import openpyxl
import math
import sys
import json
import os

# Resolve path to engine
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "../engine")))
from damage_calc_engine import DamageCalculatorEngine

EXCEL_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), "../data/HBR計算機🎭Ver.4.31.03_custom_fixed.xlsx"))

def run_tests():
    print("=== Running Exhaustive Damage Calculator Regression Tests ===")
    
    # 1. Excelから現在のアクティブな計算状態（期待値とインプット）をロード
    print(f"Loading workbook: {EXCEL_PATH} ...")
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Excel file not found: {EXCEL_PATH}")
        sys.exit(1)
        
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["ダメージ計算機"]
    
    # Load sp_mapping locally
    sp_mapping_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "../seraphdb_json/skill_sp_mapping.json"))
    sp_mapping = {}
    if os.path.exists(sp_mapping_path):
        with open(sp_mapping_path, "r", encoding="utf-8") as f:
            sp_mapping = json.load(f)

    actor_name = ws["AZ18"].value
    skill_name_active = ws["AZ20"].value
    enemy_name = ws["AZ5"].value
    
    print(f"Excel Active State - Attacker: {actor_name} | Skill: {skill_name_active} | Enemy: {enemy_name}")
    
    # Read active multipliers and stats
    excel_status_atk = ws["AJ8"].value
    excel_buff = ws["AJ67"].value
    excel_debuff = ws["AJ82"].value
    excel_crit_buff = ws["AR65"].value
    excel_resist = ws["AJ88"].value
    excel_break_rate = ws["AJ10"].value
    excel_special_effect = ws["AJ11"].value
    excel_token_ratio = ws["AS42"].value or 0.0
    
    # Stats
    stats = {
        "str": ws["AK2"].value,
        "dex": ws["AK3"].value,
        "wis": ws["AK4"].value,
        "spr": ws["AK5"].value,
        "luk": ws["AK6"].value,
        "con": ws["AK7"].value
    }
    
    engine = DamageCalculatorEngine()
    
    # Find enemy ID
    enemy_id = 13000001
    enemy = next((e for e in engine.enemies if e["name"] == enemy_name or enemy_name in e["name"]), None)
    if enemy:
        enemy_id = enemy["id"]
        
    style_id = 1010103  # 希望の暁 (手塚咲)
    style = next((s for s in engine.styles if s["id"] == style_id), None)
    weapon_type = style.get("type", "Slash") if style else "Slash"
    
    print("Starting row-by-row validation...")
    
    mismatches = 0
    passed = 0
    skipped = 0
    
    tolerance = 1e-4
    
    for r in range(12, 571):
        sp_val = ws.cell(row=r, column=2).value # column B
        c_val = ws.cell(row=r, column=3).value  # column C
        
        # Filter headers and empty cells
        if not c_val or c_val == "スキル名" or sp_val == "SP消費相当":
            continue
            
        names = [n.strip() for n in str(c_val).split(",") if n.strip()]
        if not names:
            continue
            
        test_skill_name = names[0]
        
        # Read Excel values for this row
        excel_base_dmg_normal = ws.cell(row=r, column=23).value  # W column
        excel_base_dmg_crit = ws.cell(row=r, column=24).value    # X column
        excel_expected_normal = ws.cell(row=r, column=25).value  # Y column
        excel_expected_crit = ws.cell(row=r, column=28).value    # AB column
        
        if excel_base_dmg_normal is None or excel_base_dmg_crit is None:
            skipped += 1
            continue
            
        # Determine if this row is the active skill's row
        is_active_skill_row = (skill_name_active in names)
        
        # Determine active zone from Excel state
        element_val = ws["AZ47"].value
        field_val = ws["AS32"].value
        active_zone = "None"
        try:
            field_f = float(field_val) if field_val is not None else 0.0
        except ValueError:
            field_f = 0.0

        if field_f > 0.0 and element_val and element_val != "無":
            element_map = {
                "火": "FireZone",
                "氷": "IceZone",
                "雷": "ThunderZone",
                "光": "LightZone",
                "闇": "DarkZone"
            }
            active_zone = element_map.get(element_val, "None")

        # Look up mapping info
        mapping_info = None
        if test_skill_name:
            clean_name = test_skill_name.replace("[単独発動]", "").split("[")[0].split("(")[0].split("（")[0].strip()
            mapping_info = sp_mapping.get(test_skill_name)
            if not mapping_info:
                mapping_info = sp_mapping.get(clean_name)

        sp_override = None
        e_mapped_override = None
        is_aoe_override = None
        is_normal_attack_override = None
        is_pursuit_override = None

        if mapping_info:
            sp_val = mapping_info.get("sp")
            if sp_val is not None and sp_val != "-":
                sp_override = float(sp_val)
            else:
                sp_override = 0.0

            e_val = mapping_info.get("e")
            if e_val is not None:
                e_mapped_override = float(e_val)

            is_aoe_override = bool(mapping_info.get("is_aoe", False))
            is_normal_attack_override = bool(mapping_info.get("is_normal_attack", False))
            is_pursuit_override = bool(mapping_info.get("is_pursuit", False))
        else:
            if test_skill_name:
                if "通常攻撃" in test_skill_name:
                    is_normal_attack_override = True
                elif "追撃" in test_skill_name:
                    is_pursuit_override = True

        # Build input
        input_data = {
            "attacker": {
                "characterId": actor_name,
                "styleId": style_id,
                "level": 120,
                "limitBreakCount": 0,
                "stats": stats,
                "statusEffects": [
                    {"buffType": "AttackUp", "power": excel_buff * 100.0}
                ],
                "tokenRatio": excel_token_ratio
            },
            "defender": {
                "enemyId": enemy_id,
                "enemyName": enemy_name,
                "paramBorder": ws["AY5"].value,
                "destructionRate": excel_break_rate,
                "isHpTarget": (int(ws["AO11"].value or 0) == 0),
                "resistances": {
                    weapon_type: excel_resist
                },
                "statusEffects": [
                    {"debuffType": "DefenseDown", "power": excel_debuff * 100.0}
                ]
            },
            "skill": {
                "skillId": None,
                "name": test_skill_name,
                "level": float(ws["AJ12"].value or 10.0),
                "spCostOverride": sp_override,
                "eMappedOverride": e_mapped_override,
                "isAoe": is_aoe_override,
                "isNormalAttack": is_normal_attack_override,
                "isPursuit": is_pursuit_override
            },
            "activeZone": active_zone
        }
        
        try:
            result = engine.calculate_damage(input_data)
        except Exception as e:
            print(f"❌ Row {r} ({test_skill_name}) crashed: {e}")
            mismatches += 1
            continue
            
        py_base_dmg_normal = result['breakdown']['baseDamageNormal']
        py_base_dmg_crit = result['breakdown']['baseDamageCrit']
        
        row_passed = True
        
        # 1. Base damages must match for all rows
        if not math.isclose(py_base_dmg_normal, excel_base_dmg_normal, rel_tol=tolerance, abs_tol=1e-2):
            print(f"❌ Base Normal mismatch at Row {r} ({test_skill_name}): Py={py_base_dmg_normal:.4f} | Excel={excel_base_dmg_normal:.4f}")
            row_passed = False
            
        if not math.isclose(py_base_dmg_crit, excel_base_dmg_crit, rel_tol=tolerance, abs_tol=1e-2):
            print(f"❌ Base Crit mismatch at Row {r} ({test_skill_name}): Py={py_base_dmg_crit:.4f} | Excel={excel_base_dmg_crit:.4f}")
            row_passed = False
            
        # 2. Expected damages must match for the active skill row
        if is_active_skill_row:
            py_expected_normal = result['normal']['expected']
            py_expected_crit = result['critical']['expected']
            
            if not math.isclose(py_expected_normal, excel_expected_normal, rel_tol=tolerance, abs_tol=1e-2):
                print(f"❌ Expected Normal mismatch at Active Skill Row {r} ({test_skill_name}): Py={py_expected_normal:.4f} | Excel={excel_expected_normal:.4f}")
                row_passed = False
                
            if not math.isclose(py_expected_crit, excel_expected_crit, rel_tol=tolerance, abs_tol=1e-2):
                print(f"❌ Expected Crit mismatch at Active Skill Row {r} ({test_skill_name}): Py={py_expected_crit:.4f} | Excel={excel_expected_crit:.4f}")
                row_passed = False
                
            # Verify final output cells (W4 and W10) too
            excel_w4 = ws["W4"].value
            excel_w10 = ws["W10"].value
            if excel_w4 is not None and not math.isclose(py_expected_normal, excel_w4, rel_tol=tolerance, abs_tol=1e-2):
                print(f"❌ Final Expected Normal (W4) mismatch: Py={py_expected_normal:.4f} | Excel W4={excel_w4:.4f}")
                row_passed = False
            if excel_w10 is not None and not math.isclose(py_expected_crit, excel_w10, rel_tol=tolerance, abs_tol=1e-2):
                print(f"❌ Final Expected Crit (W10) mismatch: Py={py_expected_crit:.4f} | Excel W10={excel_w10:.4f}")
                row_passed = False
        
        if row_passed:
            passed += 1
        else:
            mismatches += 1

    print(f"\nValidation summary: Passed={passed} | Mismatches={mismatches} | Skipped={skipped}")
    
    if mismatches == 0:
        print("\n🎉 SUCCESS! Python damage calculator matches Excel perfectly!")
        sys.exit(0)
    else:
        print("\n😭 FAILURE! Mismatch detected between Python and Excel.")
        sys.exit(1)

if __name__ == "__main__":
    run_tests()
