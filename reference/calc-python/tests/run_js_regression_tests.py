import openpyxl
import math
import sys
import json
import os
import subprocess

EXCEL_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), "../data/HBR計算機🎭Ver.4.31.03_custom_fixed.xlsx"))
TEMP_INPUTS_PATH = os.path.join(os.path.dirname(__file__), "temp_js_inputs.json")
TEMP_OUTPUTS_PATH = os.path.join(os.path.dirname(__file__), "temp_js_outputs.json")

def run_tests():
    print("=== Running JS Damage Calculator Regression Tests ===")
    
    # 1. Excelから現在のアクティブな計算状態（期待値とインプット）をロード
    print(f"Loading workbook: {EXCEL_PATH} ...")
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Excel file not found: {EXCEL_PATH}")
        sys.exit(1)
        
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["ダメージ計算機"]
    
    actor_name = ws["AZ18"].value
    skill_name_active = ws["AZ20"].value
    enemy_name = ws["AZ5"].value
    
    print(f"Excel Active State - Attacker: {actor_name} | Skill: {skill_name_active} | Enemy: {enemy_name}")
    
    # Read active multipliers and stats
    excel_status_atk = ws["AJ8"].value
    excel_buff = ws["AJ67"].value
    excel_debuff = ws["AJ82"].value
    excel_crit_buff = ws["AR65"].value
    excel_resist = ws["AJ88"].value
    excel_break_rate = ws["AJ10"].value
    excel_special_effect = ws["AJ11"].value
    excel_token_ratio = ws["AS42"].value or 0.0
    
    # Stats
    stats = {
        "str": ws["AK2"].value,
        "dex": ws["AK3"].value,
        "wis": ws["AK4"].value,
        "spr": ws["AK5"].value,
        "luk": ws["AK6"].value,
        "con": ws["AK7"].value
    }
    
    # Load Master Data for mappings
    enemies_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "../json/enemies.json"))
    styles_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "../json/styles.json"))
    if not os.path.exists(enemies_path) or not os.path.exists(styles_path):
        print("❌ Symbolic links or master json files are missing!")
        sys.exit(1)
        
    with open(enemies_path, "r", encoding="utf-8") as f:
        enemies = json.load(f)
    with open(styles_path, "r", encoding="utf-8") as f:
        styles = json.load(f)
        
    # Find enemy ID
    enemy_id = 13000001
    enemy = next((e for e in enemies if e["name"] == enemy_name or enemy_name in e["name"]), None)
    if enemy:
        enemy_id = enemy["id"]
        
    style_id = 1010103  # 希望の暁 (手塚咲)
    style = next((s for s in styles if s["id"] == style_id), None)
    weapon_type = style.get("type", "Slash") if style else "Slash"
    
    print("Preparing test case inputs from Excel rows...")
    
    test_cases = []
    
    for r in range(12, 571):
        sp_val = ws.cell(row=r, column=2).value # column B
        c_val = ws.cell(row=r, column=3).value  # column C
        
        # Filter headers and empty cells
        if not c_val or c_val == "スキル名" or sp_val == "SP消費相当":
            continue
            
        names = [n.strip() for n in str(c_val).split(",") if n.strip()]
        if not names:
            continue
            
        test_skill_name = names[0]
        
        # Read Excel values for this row
        excel_base_dmg_normal = ws.cell(row=r, column=23).value  # W column
        excel_base_dmg_crit = ws.cell(row=r, column=24).value    # X column
        excel_expected_normal = ws.cell(row=r, column=25).value  # Y column
        excel_expected_crit = ws.cell(row=r, column=28).value    # AB column
        
        if excel_base_dmg_normal is None or excel_base_dmg_crit is None:
            continue
            
        # Determine if this row is the active skill's row
        is_active_skill_row = (skill_name_active in names)
        
        # Determine active zone from Excel state
        element_val = ws["AZ47"].value
        field_val = ws["AS32"].value
        active_zone = "None"
        try:
            field_f = float(field_val) if field_val is not None else 0.0
        except ValueError:
            field_f = 0.0

        if field_f > 0.0 and element_val and element_val != "無":
            element_map = {
                "火": "FireZone",
                "氷": "IceZone",
                "雷": "ThunderZone",
                "光": "LightZone",
                "闇": "DarkZone"
            }
            active_zone = element_map.get(element_val, "None")

        # Build input
        input_data = {
            "attacker": {
                "characterId": actor_name,
                "styleId": style_id,
                "level": 120,
                "limitBreakCount": 0,
                "stats": stats,
                "statusEffects": [
                    {"buffType": "AttackUp", "power": excel_buff * 100.0}
                ],
                "tokenRatio": excel_token_ratio
            },
            "defender": {
                "enemyId": enemy_id,
                "enemyName": enemy_name,
                "paramBorder": ws["AY5"].value,
                "destructionRate": excel_break_rate,
                "isHpTarget": (int(ws["AO11"].value or 0) == 0),
                "resistances": {
                    weapon_type: excel_resist
                },
                "statusEffects": [
                    {"debuffType": "DefenseDown", "power": excel_debuff * 100.0}
                ]
            },
            "skill": {
                "skillId": None,
                "name": test_skill_name,
                "level": float(ws["AJ12"].value or 10.0)
            },
            "activeZone": active_zone
        }
        
        test_cases.append({
            "row": r,
            "skill_name": test_skill_name,
            "is_active_skill_row": is_active_skill_row,
            "excel_base_dmg_normal": excel_base_dmg_normal,
            "excel_base_dmg_crit": excel_base_dmg_crit,
            "excel_expected_normal": excel_expected_normal,
            "excel_expected_crit": excel_expected_crit,
            "excel_w4": ws["W4"].value,
            "excel_w10": ws["W10"].value,
            "input": input_data
        })
        
    print(f"Extracted {len(test_cases)} test cases from Excel.")
    
    # Save inputs for Node helper
    inputs_only = [tc["input"] for tc in test_cases]
    with open(TEMP_INPUTS_PATH, "w", encoding="utf-8") as f:
        json.dump(inputs_only, f)
        
    print("Running JS calculator via Node.js helper...")
    try:
        # Run node helper
        node_helper = os.path.join(os.path.dirname(__file__), "run_js_calc_helper.mjs")
        result_proc = subprocess.run(
            ["node", node_helper, TEMP_INPUTS_PATH, TEMP_OUTPUTS_PATH],
            capture_output=True,
            text=True,
            check=True
        )
    except subprocess.CalledProcessError as e:
        print("❌ Node.js helper execution failed!")
        print(f"Stdout:\n{e.stdout}")
        print(f"Stderr:\n{e.stderr}")
        clean_temp_files()
        sys.exit(1)
        
    # Read JS outputs
    if not os.path.exists(TEMP_OUTPUTS_PATH):
        print(f"❌ Output file not found: {TEMP_OUTPUTS_PATH}")
        clean_temp_files()
        sys.exit(1)
        
    with open(TEMP_OUTPUTS_PATH, "r", encoding="utf-8") as f:
        js_outputs = json.load(f)
        
    print("Validating JS outputs against Excel...")
    
    mismatches = 0
    passed = 0
    tolerance = 1e-4
    
    for i, tc in enumerate(test_cases):
        r = tc["row"]
        test_skill_name = tc["skill_name"]
        js_res = js_outputs[i]
        
        if not js_res.get("success"):
            print(f"❌ Row {r} ({test_skill_name}) crashed in JS: {js_res.get('error')}")
            mismatches += 1
            continue
            
        result = js_res["result"]
        
        js_base_dmg_normal = result['breakdown']['baseDamageNormal']
        js_base_dmg_crit = result['breakdown']['baseDamageCrit']
        
        row_passed = True
        
        # 1. Base damages must match for all rows
        if not math.isclose(js_base_dmg_normal, tc["excel_base_dmg_normal"], rel_tol=tolerance, abs_tol=1e-2):
            print(f"❌ Base Normal mismatch at Row {r} ({test_skill_name}): JS={js_base_dmg_normal:.4f} | Excel={tc['excel_base_dmg_normal']:.4f}")
            row_passed = False
            
        if not math.isclose(js_base_dmg_crit, tc["excel_base_dmg_crit"], rel_tol=tolerance, abs_tol=1e-2):
            print(f"❌ Base Crit mismatch at Row {r} ({test_skill_name}): JS={js_base_dmg_crit:.4f} | Excel={tc['excel_base_dmg_crit']:.4f}")
            row_passed = False
            
        # 2. Expected damages must match for the active skill row
        if tc["is_active_skill_row"]:
            js_expected_normal = result['normal']['expected']
            js_expected_crit = result['critical']['expected']
            
            if not math.isclose(js_expected_normal, tc["excel_expected_normal"], rel_tol=tolerance, abs_tol=1e-2):
                print(f"❌ Expected Normal mismatch at Active Skill Row {r} ({test_skill_name}): JS={js_expected_normal:.4f} | Excel={tc['excel_expected_normal']:.4f}")
                row_passed = False
                
            if not math.isclose(js_expected_crit, tc["excel_expected_crit"], rel_tol=tolerance, abs_tol=1e-2):
                print(f"❌ Expected Crit mismatch at Active Skill Row {r} ({test_skill_name}): JS={js_expected_crit:.4f} | Excel={tc['excel_expected_crit']:.4f}")
                row_passed = False
                
            # Verify final output cells (W4 and W10) too
            excel_w4 = tc["excel_w4"]
            excel_w10 = tc["excel_w10"]
            if excel_w4 is not None and not math.isclose(js_expected_normal, excel_w4, rel_tol=tolerance, abs_tol=1e-2):
                print(f"❌ Final Expected Normal (W4) mismatch: JS={js_expected_normal:.4f} | Excel W4={excel_w4:.4f}")
                row_passed = False
            if excel_w10 is not None and not math.isclose(js_expected_crit, excel_w10, rel_tol=tolerance, abs_tol=1e-2):
                print(f"❌ Final Expected Crit (W10) mismatch: JS={js_expected_crit:.4f} | Excel W10={excel_w10:.4f}")
                row_passed = False
                
        if row_passed:
            passed += 1
        else:
            mismatches += 1
            
    clean_temp_files()
    
    print(f"\nJS Validation summary: Passed={passed} | Mismatches={mismatches}")
    
    if mismatches == 0:
        print("\n🎉 SUCCESS! JS damage calculator matches Excel perfectly!")
        sys.exit(0)
    else:
        print("\n😭 FAILURE! Mismatch detected between JS and Excel.")
        sys.exit(1)

def clean_temp_files():
    if os.path.exists(TEMP_INPUTS_PATH):
        os.remove(TEMP_INPUTS_PATH)
    if os.path.exists(TEMP_OUTPUTS_PATH):
        os.remove(TEMP_OUTPUTS_PATH)

if __name__ == "__main__":
    run_tests()
